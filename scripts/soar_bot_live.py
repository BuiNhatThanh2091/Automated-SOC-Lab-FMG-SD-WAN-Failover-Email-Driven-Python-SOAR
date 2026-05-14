import json
import os
import re
import time
import imaplib
import email
from email.header import decode_header
from email.mime.text import MIMEText
import smtplib
from datetime import datetime, timedelta
from pathlib import Path

import requests
import urllib3
from netmiko import ConnectHandler

try:
    from openpyxl import Workbook, load_workbook
except Exception:
    Workbook = None
    load_workbook = None

urllib3.disable_warnings()

# =========================================================
# CONFIG
# =========================================================
FMG_URL = "https://IP_FortiManager/jsonrpc"
FMG_USER = "user"
FMG_PASS = "password"

ADOM = "root"
DEVICE_NAME = "FortiGate-80E-DR"
VDOM_NAME = "root"

FGT_SSH = {
    "device_type": "fortinet",
    "host": "IP_FortiGate-50E",
    "username": "user",
    "password": "password",
    "global_delay_factor": 2,
}

# Polling
POLL_INTERVAL_SECONDS = 3
ESCALATE_AFTER_SECONDS = 3
FOLLOWUP_MAIL_INTERVAL_SECONDS = 30

# SMTP
SMTP_ENABLED = True
SMTP_HOST = "IP_FortiMail"
SMTP_PORT = 25
SMTP_USER = "mail_user"
SMTP_PASS = "mail_password"
MAIL_TO = "target_email"

# IMAP
IMAP_ENABLED = True
IMAP_HOST = "IP_FortiMail"
IMAP_PORT = 143
IMAP_USER = "mail_user"
IMAP_PASS = "mail_password"
IMAP_USE_SSL = False

# File runtime
BASE_DIR = Path(__file__).resolve().parent
RUNTIME_DIR = BASE_DIR / "runtime"
RUNTIME_DIR.mkdir(parents=True, exist_ok=True)

STATE_FILE = RUNTIME_DIR / "state.json"
AUDIT_FILE = RUNTIME_DIR / "audit.txt"
MAIL_LOG_FILE = RUNTIME_DIR / "mail_outbox.txt"
DEBUG_FILE = RUNTIME_DIR / "debug.txt"
REPORT_TXT_FILE = RUNTIME_DIR / "report.txt"
REPORT_XLSX_FILE = RUNTIME_DIR / "report.xlsx"

REPORT_FIELDS = [
    "ticket_id",
    "device_name",
    "ip",
    "conn_status",
    "conf_status",
    "pkg_name",
    "pkg_status",
    "admin_user",
    "source_ip",
    "policy_id",
    "action_type",
    "current_state",
    "approval_result",
    "manual_guide_sent",
    "first_seen",
    "last_seen",
    "last_mail_sent",
    "final_result",
]


# =========================================================
# HELPERS
# =========================================================
def now_ts() -> str:
    return datetime.now().isoformat(timespec="seconds")


def parse_ts(ts: str | None):
    if not ts:
        return None
    try:
        return datetime.fromisoformat(ts)
    except Exception:
        return None


def log_debug(msg: str):
    with open(DEBUG_FILE, "a", encoding="utf-8") as f:
        f.write(f"[{now_ts()}] {msg}\n")


def log_audit(ticket_id: str, event: str, detail: str):
    line = {
        "ts": now_ts(),
        "ticket": ticket_id,
        "device": DEVICE_NAME,
        "event": event,
        "detail": detail,
    }
    with open(AUDIT_FILE, "a", encoding="utf-8") as f:
        f.write(json.dumps(line, ensure_ascii=False) + "\n")


def ensure_files():
    if not STATE_FILE.exists():
        save_state({
            "device_name": DEVICE_NAME,
            "ip": None,
            "current_state": "NORMAL",
            "incident_open": False,
            "ticket_id": None,
            "first_seen": None,
            "last_seen": None,
            "last_mail_sent": None,
            "mail_count": 0,
            "approval_result": None,
            "manual_guide_sent": False,
            "manual_guide_sent_at": None,
            "it_execution_confirmed": False,
            "policy_id": None,
            "admin_user": None,
            "source_ip": None,
            "action_type": None,
            "conf_status": None,
            "conn_status": None,
            "pkg_name": None,
            "pkg_status": None,
            "final_result": None,
            "last_reply_uid": None,
        })

    for file_path in [AUDIT_FILE, MAIL_LOG_FILE, DEBUG_FILE, REPORT_TXT_FILE]:
        if not file_path.exists():
            file_path.write_text("", encoding="utf-8")


def load_state() -> dict:
    ensure_files()
    try:
        with open(STATE_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception as e:
        log_debug(f"load_state failed: {e}")
        return {
            "device_name": DEVICE_NAME,
            "ip": None,
            "current_state": "NORMAL",
            "incident_open": False,
            "ticket_id": None,
        }


def save_state(state: dict):
    try:
        with open(STATE_FILE, "w", encoding="utf-8") as f:
            json.dump(state, f, indent=2, ensure_ascii=False)
    except Exception as e:
        log_debug(f"save_state failed: {e}")


def build_report_record(state: dict) -> dict:
    return {
        "ticket_id": state.get("ticket_id"),
        "device_name": state.get("device_name", DEVICE_NAME),
        "ip": state.get("ip"),
        "conn_status": state.get("conn_status"),
        "conf_status": state.get("conf_status"),
        "pkg_name": state.get("pkg_name"),
        "pkg_status": state.get("pkg_status"),
        "admin_user": state.get("admin_user"),
        "source_ip": state.get("source_ip"),
        "policy_id": state.get("policy_id"),
        "action_type": state.get("action_type"),
        "current_state": state.get("current_state"),
        "approval_result": state.get("approval_result"),
        "manual_guide_sent": state.get("manual_guide_sent"),
        "first_seen": state.get("first_seen"),
        "last_seen": state.get("last_seen"),
        "last_mail_sent": state.get("last_mail_sent"),
        "final_result": state.get("final_result"),
    }


def append_report_txt(record: dict):
    try:
        write_header = (not REPORT_TXT_FILE.exists()) or REPORT_TXT_FILE.stat().st_size == 0
        with open(REPORT_TXT_FILE, "a", encoding="utf-8") as f:
            if write_header:
                f.write("\t".join(REPORT_FIELDS) + "\n")
            values = ["" if record.get(field) is None else str(record.get(field)) for field in REPORT_FIELDS]
            f.write("\t".join(values) + "\n")
    except Exception as e:
        log_debug(f"append_report_txt failed: {e}")


def append_report_xlsx(record: dict):
    if Workbook is None or load_workbook is None:
        return
    try:
        if REPORT_XLSX_FILE.exists():
            wb = load_workbook(REPORT_XLSX_FILE)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "report"
            ws.append(REPORT_FIELDS)

        ws.append([record.get(field) for field in REPORT_FIELDS])
        wb.save(REPORT_XLSX_FILE)
    except Exception as e:
        log_debug(f"append_report_xlsx failed: {e}")


def write_reports_per_cycle(state: dict):
    record = build_report_record(state)
    append_report_txt(record)
    append_report_xlsx(record)


def normalize_conf_status(value):
    text = str(value).strip().lower()
    if text in {"insync", "in-sync"}:
        return "in-sync"
    if text in {"outsync", "out-of-sync", "out of sync"}:
        return "out-of-sync"
    if text in {"modified", "mod"}:
        return "modified"
    return text


def normalize_conn_status(value):
    text = str(value).strip().lower()
    if text in {"up", "1"}:
        return "up"
    if text in {"down", "0", "unknown"}:
        return "down"
    return text


def normalize_pkg_status(value):
    text = str(value).strip().lower()
    if text in {"in sync", "in-sync", "insync", "ok", "installed"}:
        return "in-sync"
    if text in {"out of sync", "out-of-sync", "conflict", "modified", "mismatch", "warning", "error"}:
        return "out-of-sync"
    return text


def html_to_text(raw_html: str) -> str:
    text = re.sub(r"(?is)<(script|style).*?>.*?(</\1>)", "", raw_html)
    text = re.sub(r"(?s)<.*?>", " ", text)
    text = text.replace("&nbsp;", " ").replace("&lt;", "<").replace("&gt;", ">").replace("&amp;", "&")
    text = re.sub(r"\s+", " ", text).strip()
    return text


def parse_yes_no(text: str) -> str:
    clean = text.strip().upper()
    tokens = list(re.finditer(r"\b(YES|NO)\b", clean))
    if not tokens:
        return "UNKNOWN"

    # Use the first explicit token in the newest reply content.
    first = tokens[0].group(1)
    if first in {"YES", "NO"}:
        return first
    return "UNKNOWN"


def extract_latest_reply_text(subject: str, body: str) -> str:
    # Focus on the newly typed part and ignore quoted thread history.
    lines = (f"{subject}\n{body}").splitlines()
    collected = []
    for line in lines:
        striped = line.strip()

        if not striped:
            if collected:
                break
            continue

        lower = striped.lower()
        if striped.startswith(">"):
            break
        if lower.startswith("on ") and " wrote:" in lower:
            break
        if lower.startswith("from:") or lower.startswith("subject:"):
            break
        if lower.startswith("-----original message-----"):
            break

        collected.append(striped)
        if len(" ".join(collected)) > 300:
            break

    return "\n".join(collected).strip()


# =========================================================
# FMG API
# =========================================================
def fmg_login() -> str:
    payload = {
        "method": "exec",
        "params": [{
            "url": "/sys/login/user",
            "data": {"user": FMG_USER, "passwd": FMG_PASS}
        }],
        "id": 1
    }
    res = requests.post(FMG_URL, json=payload, verify=False, timeout=20).json()
    session = res.get("session")
    if not session:
        raise RuntimeError(f"FMG login failed: {res}")
    return session


def fmg_logout(session: str):
    payload = {
        "method": "exec",
        "params": [{"url": "/sys/logout"}],
        "session": session,
        "id": 99
    }
    requests.post(FMG_URL, json=payload, verify=False, timeout=15)


def fmg_get(session: str, url: str, fields=None):
    param = {"url": url}
    if fields:
        param["fields"] = fields

    payload = {
        "method": "get",
        "params": [param],
        "session": session,
        "id": 2,
        "verbose": 1
    }
    res = requests.post(FMG_URL, json=payload, verify=False, timeout=20).json()

    result = res.get("result", [])
    if not result:
        raise RuntimeError(f"No result for URL {url}: {res}")

    status = result[0].get("status", {})
    if status.get("code") != 0:
        raise RuntimeError(f"FMG get failed for {url}: {res}")

    return result[0].get("data", {})


def get_phase1_status():
    session = None
    try:
        session = fmg_login()

        device_data = fmg_get(
            session,
            f"/dvmdb/adom/{ADOM}/device/{DEVICE_NAME}",
            fields=["name", "ip", "hostname", "conf_status", "conn_status", "db_status", "vdom"]
        )

        pkg_data = fmg_get(
            session,
            f"/pm/config/adom/{ADOM}/_package/status/{DEVICE_NAME}/{VDOM_NAME}"
        )

        result = {
            "device_name": device_data.get("name", DEVICE_NAME),
            "ip": device_data.get("ip", "unknown"),
            "host_name": device_data.get("hostname", "unknown"),
            "conf_status": normalize_conf_status(device_data.get("conf_status")),
            "conn_status": normalize_conn_status(device_data.get("conn_status")),
            "db_status": str(device_data.get("db_status", "unknown")).lower(),
            "pkg_name": pkg_data.get("pkg", "unknown"),
            "pkg_status": normalize_pkg_status(pkg_data.get("status")),
            "check_time": now_ts(),
        }
        return result

    finally:
        if session:
            fmg_logout(session)


# =========================================================
# PHASE 2 - INVESTIGATION
# =========================================================
def get_recent_event_logs():
    net_connect = ConnectHandler(**FGT_SSH)
    net_connect.send_command("execute log filter reset")
    net_connect.send_command("execute log filter category event")
    net_connect.send_command("execute log filter view-lines 20")
    output = net_connect.send_command("execute log display")
    net_connect.disconnect()
    return output


def split_log_entries(output: str):
    lines = output.splitlines()
    entries = []
    current = []

    for line in lines:
        line = line.strip()
        if re.match(r"^\d+:\s+date=", line):
            if current:
                entries.append(" ".join(current))
                current = []
            current.append(line)
        elif current:
            current.append(line)

    if current:
        entries.append(" ".join(current))
    return entries


def pick_relevant_entry(entries):
    keywords = [
        "Configuration changed",
        "firewall.policy",
        "policy",
        "delete",
        "edit",
        "create",
        "remove",
        "update",
    ]
    for entry in entries:
        for kw in keywords:
            if kw.lower() in entry.lower():
                return entry
    return entries[0] if entries else ""


def detect_action(log_text: str):
    text = log_text.lower()
    if any(x in text for x in ["delete", "remove"]):
        return "delete"
    if any(x in text for x in ["edit", "update", "modify", "change"]):
        return "edit"
    if any(x in text for x in ["create", "add", "new"]):
        return "create"
    return "unknown"


def extract_policy_id(log_text: str):
    patterns = [
        r"firewall\.policy\s+(\d+)",
        r"policy(?:\s+id)?\s+(\d+)",
        r"delete\s+(\d+)",
        r"edit\s+(\d+)",
    ]
    for pat in patterns:
        m = re.search(pat, log_text, re.IGNORECASE)
        if m:
            return m.group(1)
    return None


def investigate_change():
    try:
        output = get_recent_event_logs()
        entries = split_log_entries(output)
        entry = pick_relevant_entry(entries)
        if not entry:
            return {
                "admin_user": None,
                "source_ip": None,
                "policy_id": None,
                "action_type": "unknown",
                "raw_message": "no relevant log found",
            }

        user_match = re.search(r'user="([^"]+)"', entry)
        ip_match = re.search(r'(?:srcip=|ui="(?:https|ssh)\()([0-9]+\.[0-9]+\.[0-9]+\.[0-9]+)', entry)
        msg_match = re.search(r'msg="([^"]+)"', entry)

        return {
            "admin_user": user_match.group(1) if user_match else None,
            "source_ip": ip_match.group(1) if ip_match else None,
            "policy_id": extract_policy_id(entry),
            "action_type": detect_action(entry),
            "raw_message": msg_match.group(1) if msg_match else entry,
        }
    except Exception as e:
        log_debug(f"investigate_change failed: {e}")
        return {
            "admin_user": None,
            "source_ip": None,
            "policy_id": None,
            "action_type": "unknown",
            "raw_message": f"investigation failed: {e}",
        }


# =========================================================
# SMTP / IMAP
# =========================================================
def send_mail(subject: str, body: str):
    with open(MAIL_LOG_FILE, "a", encoding="utf-8") as f:
        f.write(f"[{now_ts()}]\nSUBJECT: {subject}\n{body}\n{'-'*80}\n")

    if not SMTP_ENABLED:
        return True

    try:
        msg = MIMEText(body, _charset="utf-8")
        msg["Subject"] = subject
        msg["From"] = SMTP_USER
        msg["To"] = MAIL_TO

        server = smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=15)
        server.ehlo()
        try:
            server.login(SMTP_USER, SMTP_PASS)
        except smtplib.SMTPNotSupportedError:
            # Some internal relay servers allow sending without AUTH.
            pass
        server.send_message(msg)
        server.quit()
        print(f"[MAIL] Sent: {subject}")
        return True
    except Exception as e:
        log_debug(f"send_mail failed: {e}")
        print(f"[MAIL][ERROR] Failed to send: {subject} ({e})")
        return False


def decode_mime_header(value):
    if not value:
        return ""
    parts = decode_header(value)
    decoded = []
    for part, enc in parts:
        if isinstance(part, bytes):
            decoded.append(part.decode(enc or "utf-8", errors="ignore"))
        else:
            decoded.append(part)
    return "".join(decoded)


def fetch_latest_reply_for_ticket(ticket_id: str, last_uid: str | None):
    if not IMAP_ENABLED:
        return None

    try:
        if IMAP_USE_SSL:
            mail = imaplib.IMAP4_SSL(IMAP_HOST, 993)
        else:
            mail = imaplib.IMAP4(IMAP_HOST, IMAP_PORT)

        mail.login(IMAP_USER, IMAP_PASS)
        mail.select("INBOX")

        status, data = mail.search(None, "ALL")
        if status != "OK":
            mail.logout()
            return None

        ids = data[0].split()
        ids = ids[-30:]  # check 30 mail gần nhất

        for msg_id in reversed(ids):
            uid = msg_id.decode()

            if last_uid and uid <= str(last_uid):
                continue

            status, msg_data = mail.fetch(msg_id, "(RFC822)")
            if status != "OK":
                continue

            raw_email = msg_data[0][1]
            msg = email.message_from_bytes(raw_email)

            subject = decode_mime_header(msg.get("Subject", ""))
            from_addr = decode_mime_header(msg.get("From", ""))

            body = ""
            if msg.is_multipart():
                for part in msg.walk():
                    ctype = part.get_content_type()
                    cdisp = str(part.get("Content-Disposition"))
                    if ctype in ("text/plain", "text/html") and "attachment" not in cdisp:
                        charset = part.get_content_charset() or "utf-8"
                        body = part.get_payload(decode=True).decode(charset, errors="ignore")
                        if ctype == "text/html":
                            body = html_to_text(body)
                        break
            else:
                charset = msg.get_content_charset() or "utf-8"
                body = msg.get_payload(decode=True).decode(charset, errors="ignore")
                if "<html" in body.lower() or "<div" in body.lower():
                    body = html_to_text(body)

            merged = f"{subject}\n{body}"

            if ticket_id in merged:
                latest_reply_text = extract_latest_reply_text(subject, body)
                reply = parse_yes_no(latest_reply_text)
                if reply in {"YES", "NO"}:
                    mail.logout()
                    return {
                        "uid": uid,
                        "from": from_addr,
                        "subject": subject,
                        "body": body,
                        "latest_reply_text": latest_reply_text,
                        "reply": reply,
                    }

        mail.logout()
        return None

    except Exception as e:
        log_debug(f"fetch_latest_reply_for_ticket failed: {e}")
        return None


# =========================================================
# INCIDENT WORKFLOW
# =========================================================
def is_problem(status_info: dict) -> bool:
    if status_info["conn_status"] != "up":
        return True
    if status_info["conf_status"] != "in-sync":
        return True
    if status_info["pkg_status"] != "in-sync":
        return True
    return False


def is_resolved(status_info: dict) -> bool:
    return (
        status_info["conn_status"] == "up"
        and status_info["conf_status"] == "in-sync"
        and status_info["pkg_status"] == "in-sync"
    )


def create_incident_if_needed(status_info: dict, investigation: dict):
    state = load_state()

    if state.get("incident_open"):
        state["last_seen"] = now_ts()
        state["conf_status"] = status_info["conf_status"]
        state["conn_status"] = status_info["conn_status"]
        state["pkg_name"] = status_info["pkg_name"]
        state["pkg_status"] = status_info["pkg_status"]
        state["ip"] = status_info.get("ip")

        if investigation.get("admin_user"):
            state["admin_user"] = investigation.get("admin_user")
        if investigation.get("source_ip"):
            state["source_ip"] = investigation.get("source_ip")
        if investigation.get("policy_id"):
            state["policy_id"] = investigation.get("policy_id")
        if investigation.get("action_type"):
            state["action_type"] = investigation.get("action_type")

        save_state(state)
        log_audit(
            state.get("ticket_id", "UNKNOWN"),
            "RECHECK",
            f"conf={status_info['conf_status']}, conn={status_info['conn_status']}, pkg={status_info['pkg_status']}",
        )
        return state

    ticket_id = f"TICKET-{datetime.now().strftime('%Y%m%d-%H%M%S')}"

    state.update({
        "incident_open": True,
        "ticket_id": ticket_id,
        "ip": status_info.get("ip"),
        "current_state": "DETECTED",
        "first_seen": now_ts(),
        "last_seen": now_ts(),
        "last_mail_sent": None,
        "mail_count": 0,
        "approval_result": None,
        "manual_guide_sent": False,
        "manual_guide_sent_at": None,
        "it_execution_confirmed": False,
        "policy_id": investigation.get("policy_id"),
        "admin_user": investigation.get("admin_user"),
        "source_ip": investigation.get("source_ip"),
        "action_type": investigation.get("action_type"),
        "conf_status": status_info["conf_status"],
        "conn_status": status_info["conn_status"],
        "pkg_name": status_info["pkg_name"],
        "pkg_status": status_info["pkg_status"],
        "final_result": None,
    })
    save_state(state)
    log_audit(ticket_id, "DETECTED", f"conf={status_info['conf_status']}, conn={status_info['conn_status']}, pkg={status_info['pkg_status']}")
    return state


def send_initial_alert_if_needed():
    state = load_state()
    if not state.get("incident_open"):
        return

    if state.get("mail_count", 0) > 0:
        return

    subject = f"[ALERT] {DEVICE_NAME} - out of sync detected"
    body = (
        f"Ticket ID: {state['ticket_id']}\n"
        f"Device: {DEVICE_NAME}\n"
        f"Connection Status: {state['conn_status']}\n"
        f"Config Status: {state['conf_status']}\n"
        f"Policy Package: {state['pkg_name']}\n"
        f"Policy Package Status: {state['pkg_status']}\n"
        f"Admin User: {state.get('admin_user')}\n"
        f"Source IP: {state.get('source_ip')}\n"
        f"Policy ID: {state.get('policy_id')}\n"
        f"Action Type: {state.get('action_type')}\n\n"
        f"Bot sẽ tiếp tục polling. Nếu sự cố kéo dài, bot sẽ gửi mail yêu cầu phản hồi YES/NO."
    )

    if send_mail(subject, body):
        state["current_state"] = "ALERT_SENT_LV1"
        state["mail_count"] += 1
        state["last_mail_sent"] = now_ts()
        save_state(state)
        log_audit(state["ticket_id"], "MAIL_SENT_LV1", "initial alert sent")


def send_escalation_if_needed():
    state = load_state()
    if not state.get("incident_open"):
        return

    if state.get("current_state") != "ALERT_SENT_LV1":
        return

    first_seen = parse_ts(state.get("first_seen"))
    if not first_seen:
        return

    if datetime.now() - first_seen < timedelta(seconds=ESCALATE_AFTER_SECONDS):
        return

    subject = f"[APPROVAL REQUIRED] {DEVICE_NAME} - reply YES or NO"
    body = (
        f"Ticket ID: {state['ticket_id']}\n"
        f"Sự cố vẫn còn tồn tại sau {ESCALATE_AFTER_SECONDS} giây.\n"
        f"Device: {DEVICE_NAME}\n"
        f"Connection Status: {state['conn_status']}\n"
        f"Config Status: {state['conf_status']}\n"
        f"Policy Package: {state['pkg_name']}\n"
        f"Policy Package Status: {state['pkg_status']}\n\n"
        f"Hãy reply mail này bằng YES hoặc NO.\n"
        f"YES = Bot sẽ gửi hướng dẫn xử lý thủ công.\n"
        f"NO  = Bot sẽ yêu cầu đồng bộ lại với FortiManager và tiếp tục polling."
    )

    if send_mail(subject, body):
        state["current_state"] = "WAITING_APPROVAL"
        state["mail_count"] += 1
        state["last_mail_sent"] = now_ts()
        save_state(state)
        log_audit(state["ticket_id"], "MAIL_SENT_LV2", "approval request sent")


def process_mail_reply_if_any():
    state = load_state()
    if not state.get("incident_open"):
        return

    if state.get("current_state") not in {
        "DETECTED",
        "ALERT_SENT_LV1",
        "WAITING_APPROVAL",
        "MANUAL_ROLLBACK_GUIDE_SENT",
        "SYNC_REQUESTED",
        "NO_ROLLBACK_APPROVED",
    }:
        return

    ticket_id = state["ticket_id"]
    last_uid = state.get("last_reply_uid")

    reply_mail = fetch_latest_reply_for_ticket(ticket_id, last_uid)
    if not reply_mail:
        return

    state["last_reply_uid"] = reply_mail["uid"]
    reply = reply_mail["reply"]

    if state.get("current_state") == "WAITING_APPROVAL" and reply == "YES" and not state.get("manual_guide_sent"):
        cli_hint = ""
        if state.get("policy_id"):
            cli_hint = f"config firewall policy\n    delete {state['policy_id']}\nend"
        else:
            cli_hint = "Không bóc được policy ID từ log. IT cần kiểm tra và xử lý thủ công trên FMG/FGT."

        state["approval_result"] = "YES"
        state["current_state"] = "APPROVED_MANUAL_ROLLBACK"
        save_state(state)
        log_audit(ticket_id, "MANUAL_ROLLBACK_APPROVED", "IT approved manual rollback")

        subject = f"[MANUAL REMEDIATION GUIDE] {DEVICE_NAME}"
        body = (
            f"Ticket ID: {ticket_id}\n"
            f"Device: {DEVICE_NAME}\n"
            f"Device IP: {state.get('ip')}\n"
            f"First Seen: {state.get('first_seen')}\n"
            f"Current State: {state.get('current_state')}\n"
            f"Connection Status: {state.get('conn_status')}\n"
            f"Config Status: {state.get('conf_status')}\n"
            f"Policy Package: {state.get('pkg_name')}\n"
            f"Policy Package Status: {state.get('pkg_status')}\n"
            f"Admin User: {state.get('admin_user')}\n"
            f"Source IP: {state.get('source_ip')}\n"
            f"Policy ID: {state.get('policy_id')}\n"
            f"Action Type: {state.get('action_type')}\n\n"
            f"IT đã phản hồi YES.\n"
            f"Bot không tự rollback.\n"
            f"Đề nghị IT thực hiện xử lý thủ công như sau (human-in-the-loop):\n\n"
            f"{cli_hint}\n\n"
            f"Sau khi làm xong, bot sẽ tiếp tục polling để xác minh trạng thái."
        )
        if send_mail(subject, body):
            state["current_state"] = "MANUAL_ROLLBACK_GUIDE_SENT"
            state["manual_guide_sent"] = True
            state["manual_guide_sent_at"] = now_ts()
            state["mail_count"] += 1
            state["last_mail_sent"] = now_ts()
            save_state(state)
            log_audit(ticket_id, "MAIL_REPLY_YES", "manual remediation guide sent")
            log_audit(ticket_id, "MANUAL_ROLLBACK_GUIDE_SENT", "guidance sent after YES")

    elif state.get("current_state") == "WAITING_APPROVAL" and reply == "NO" and state.get("current_state") != "SYNC_REQUESTED":
        subject = f"[SYNC REQUIRED] {DEVICE_NAME}"
        body = (
            f"Ticket ID: {ticket_id}\n"
            f"IT đã phản hồi NO.\n"
            f"Không rollback.\n"
            f"Đề nghị IT đồng bộ lại với FortiManager hoặc xác nhận baseline hợp lệ.\n"
            f"Bot sẽ tiếp tục polling cho đến khi trạng thái trở lại bình thường."
        )
        if send_mail(subject, body):
            state["approval_result"] = "NO"
            state["current_state"] = "SYNC_REQUESTED"
            state["mail_count"] += 1
            state["last_mail_sent"] = now_ts()
            save_state(state)
            log_audit(ticket_id, "MAIL_REPLY_NO", "sync requested after NO")
            log_audit(ticket_id, "SYNC_REQUESTED", "sync request sent after NO")

    elif state.get("current_state") == "MANUAL_ROLLBACK_GUIDE_SENT" and reply == "YES":
        cli_hint = ""
        if state.get("policy_id"):
            cli_hint = f"config firewall policy\n    delete {state['policy_id']}\nend"
        else:
            cli_hint = "Không bóc được policy ID từ log. IT cần kiểm tra và xử lý thủ công trên FMG/FGT."

        subject = f"[MANUAL REMEDIATION REMINDER] {DEVICE_NAME}"
        body = (
            f"Ticket ID: {ticket_id}\n"
            f"Bot đã nhận thêm phản hồi YES.\n"
            f"Gửi lại hướng dẫn xử lý thủ công để IT tiện thao tác:\n\n"
            f"{cli_hint}\n\n"
            f"Sau khi xử lý xong, bot sẽ tiếp tục polling để xác minh trạng thái."
        )
        if send_mail(subject, body):
            state["mail_count"] += 1
            state["last_mail_sent"] = now_ts()
            save_state(state)
            log_audit(ticket_id, "MAIL_REPLY_YES", "manual guide reminder sent")
            log_audit(ticket_id, "MANUAL_ROLLBACK_GUIDE_SENT", "manual guide reminder re-sent")

    elif state.get("current_state") == "MANUAL_ROLLBACK_GUIDE_SENT" and reply == "NO":
        subject = f"[NO ROLLBACK APPROVED] {DEVICE_NAME}"
        body = (
            f"Ticket ID: {ticket_id}\n"
            f"IT đã phản hồi NO sau khi nhận hướng dẫn rollback.\n"
            f"Bot ghi nhận quyết định không rollback tại thời điểm này.\n"
            f"Trạng thái sẽ được theo dõi tiếp qua polling liên tục.\n"
            f"Nếu đồng bộ trở lại, bot sẽ tự động ghi RESOLVED/CLOSED."
        )
        if send_mail(subject, body):
            state["approval_result"] = "NO"
            state["current_state"] = "NO_ROLLBACK_APPROVED"
            state["mail_count"] += 1
            state["last_mail_sent"] = now_ts()
            save_state(state)
            log_audit(ticket_id, "MAIL_REPLY_NO", "NO received after manual guide")
            log_audit(ticket_id, "NO_ROLLBACK_APPROVED", "IT approved no rollback")

    elif state.get("current_state") == "SYNC_REQUESTED" and reply == "NO":
        subject = f"[SYNC REQUIRED - REMINDER] {DEVICE_NAME}"
        body = (
            f"Ticket ID: {ticket_id}\n"
            f"Bot đã nhận thêm phản hồi NO.\n"
            f"Đề nghị IT đồng bộ lại với FortiManager hoặc xác nhận baseline hợp lệ.\n"
            f"Bot vẫn tiếp tục polling cho đến khi trạng thái trở lại bình thường."
        )
        if send_mail(subject, body):
            state["mail_count"] += 1
            state["last_mail_sent"] = now_ts()
            save_state(state)
            log_audit(ticket_id, "MAIL_REPLY_NO", "sync reminder sent")
            log_audit(ticket_id, "SYNC_REQUESTED", "sync reminder re-sent")

    else:
        print("[INFO] Reply detected but ignored because incident state does not match the reply workflow.")


def send_followup_if_needed():
    state = load_state()
    if not state.get("incident_open"):
        return

    if state.get("current_state") not in {
        "MANUAL_ROLLBACK_GUIDE_SENT",
        "SYNC_REQUESTED",
        "NO_ROLLBACK_APPROVED",
    }:
        return

    last_mail_ts = parse_ts(state.get("last_mail_sent"))
    if not last_mail_ts:
        return

    if datetime.now() - last_mail_ts < timedelta(seconds=FOLLOWUP_MAIL_INTERVAL_SECONDS):
        return

    ticket_id = state.get("ticket_id")
    if not ticket_id:
        return

    if state.get("current_state") == "MANUAL_ROLLBACK_GUIDE_SENT":
        subject = f"[FOLLOW-UP] {DEVICE_NAME} - waiting IT manual action"
        body = (
            f"Ticket ID: {ticket_id}\n"
            f"Thiết bị vẫn chưa đồng bộ lại.\n"
            f"Current Status: conn={state.get('conn_status')} conf={state.get('conf_status')} pkg={state.get('pkg_status')}\n"
            f"Bot đang chờ IT thực hiện xử lý thủ công theo hướng dẫn đã gửi.\n"
            f"Sau khi xử lý xong, bot sẽ tự kiểm tra lại trạng thái."
        )
        if send_mail(subject, body):
            state["mail_count"] += 1
            state["last_mail_sent"] = now_ts()
            save_state(state)
            log_audit(ticket_id, "WAITING_IT_EXECUTION", "follow-up reminder sent")

    elif state.get("current_state") == "SYNC_REQUESTED":
        subject = f"[FOLLOW-UP] {DEVICE_NAME} - sync still required"
        body = (
            f"Ticket ID: {ticket_id}\n"
            f"Thiết bị vẫn chưa đồng bộ lại với FortiManager.\n"
            f"Current Status: conn={state.get('conn_status')} conf={state.get('conf_status')} pkg={state.get('pkg_status')}\n"
            f"Đề nghị IT thực hiện sync và bot sẽ tiếp tục xác minh tự động."
        )
        if send_mail(subject, body):
            state["mail_count"] += 1
            state["last_mail_sent"] = now_ts()
            save_state(state)
            log_audit(ticket_id, "SYNC_REQUESTED", "follow-up sync reminder sent")

    elif state.get("current_state") == "NO_ROLLBACK_APPROVED":
        subject = f"[FOLLOW-UP] {DEVICE_NAME} - no rollback approved"
        body = (
            f"Ticket ID: {ticket_id}\n"
            f"Bot đã ghi nhận quyết định NO ROLLBACK.\n"
            f"Trạng thái hiện tại vẫn chưa về normal: conn={state.get('conn_status')} conf={state.get('conf_status')} pkg={state.get('pkg_status')}\n"
            f"Bot tiếp tục monitor và sẽ tự đóng incident khi trạng thái bình thường trở lại."
        )
        if send_mail(subject, body):
            state["mail_count"] += 1
            state["last_mail_sent"] = now_ts()
            save_state(state)
            log_audit(ticket_id, "BASELINE_ACCEPTED", "follow-up no-rollback monitoring mail sent")


def close_incident_if_resolved(status_info: dict):
    state = load_state()
    if not state.get("incident_open"):
        return

    if not is_resolved(status_info):
        return

    ticket_id = state["ticket_id"]
    state["current_state"] = "RESOLVED"
    state["final_result"] = "resolved"
    state["last_seen"] = now_ts()
    save_state(state)
    log_audit(ticket_id, "RESOLVED", "device returned to normal state")

    state.update({
        "incident_open": False,
        "ticket_id": None,
        "current_state": "NORMAL",
        "approval_result": None,
        "manual_guide_sent": False,
        "manual_guide_sent_at": None,
        "it_execution_confirmed": False,
        "ip": status_info.get("ip"),
        "policy_id": None,
        "admin_user": None,
        "source_ip": None,
        "action_type": None,
        "conf_status": status_info["conf_status"],
        "conn_status": status_info["conn_status"],
        "pkg_name": status_info["pkg_name"],
        "pkg_status": status_info["pkg_status"],
        "first_seen": None,
        "last_seen": now_ts(),
        "last_mail_sent": None,
        "mail_count": 0,
        "final_result": "resolved",
        "last_reply_uid": state.get("last_reply_uid"),
    })
    save_state(state)
    log_audit(ticket_id, "CLOSED", "incident closed")


# =========================================================
# MAIN LOOP
# =========================================================
def process_cycle():
    status_info = get_phase1_status()
    state = load_state()

    state["ip"] = status_info.get("ip")
    state["conf_status"] = status_info["conf_status"]
    state["conn_status"] = status_info["conn_status"]
    state["pkg_name"] = status_info["pkg_name"]
    state["pkg_status"] = status_info["pkg_status"]
    state["last_seen"] = now_ts()
    save_state(state)

    if is_problem(status_info):
        investigation = investigate_change()
        create_incident_if_needed(status_info, investigation)
        send_initial_alert_if_needed()
        send_escalation_if_needed()
        process_mail_reply_if_any()
        send_followup_if_needed()
    else:
        close_incident_if_resolved(status_info)

    state = load_state()
    print(
        f"[{now_ts()}] conn={state.get('conn_status')} conf={state.get('conf_status')} "
        f"pkg={state.get('pkg_status')} state={state.get('current_state')}"
    )
    write_reports_per_cycle(state)


def main():
    ensure_files()
    print(f"[*] Bot đang polling mỗi {POLL_INTERVAL_SECONDS} giây. Nhấn Ctrl+C để dừng.")
    try:
        while True:
            try:
                process_cycle()
            except Exception as e:
                log_debug(f"process_cycle failed: {e}")
                print(f"[ERROR] {e}")
            time.sleep(POLL_INTERVAL_SECONDS)
    except KeyboardInterrupt:
        print("\n[STOP] Dừng bot.")


if __name__ == "__main__":
    main()